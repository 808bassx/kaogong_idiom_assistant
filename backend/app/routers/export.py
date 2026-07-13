"""
数据导出 API 路由
支持 CSV、Excel、Markdown、JSON 格式
"""
import csv
import json
import io
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.database import get_db
from app.models import Word

router = APIRouter(prefix="/api/export", tags=["数据导出"])


async def _get_words(db: AsyncSession, word_ids: list[int] = None) -> list[dict]:
    """获取要导出的词语列表"""
    query = select(Word)
    if word_ids:
        query = query.where(Word.id.in_(word_ids))
    query = query.order_by(Word.word)
    result = await db.execute(query)
    return [w.to_dict() for w in result.scalars().all()]


@router.get("/csv")
async def export_csv(
    word_ids: str = Query(None, description="逗号分隔的词语ID"),
    db: AsyncSession = Depends(get_db),
):
    """导出为 CSV"""
    ids = [int(x) for x in word_ids.split(",")] if word_ids else None
    words = await _get_words(db, ids)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["词语", "拼音", "释义", "出处", "用法", "例句",
                      "近义词", "反义词", "易混词", "记忆技巧", "标签",
                      "掌握状态", "复习次数", "错误次数", "收藏"])

    for w in words:
        writer.writerow([
            w["word"], w["pinyin"], w["meaning"], w["source"],
            w["usage"], w["example"], w["synonym"], w["antonym"],
            w["confusable"], w["memory_tip"],
            ",".join(w["tags"]) if isinstance(w["tags"], list) else w["tags"],
            "是" if w["is_mastered"] else "否",
            w["review_count"], w["error_count"],
            "是" if w["is_favorite"] else "否",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=idioms.csv"},
    )


@router.get("/json")
async def export_json(
    word_ids: str = Query(None, description="逗号分隔的词语ID"),
    db: AsyncSession = Depends(get_db),
):
    """导出为 JSON"""
    ids = [int(x) for x in word_ids.split(",")] if word_ids else None
    words = await _get_words(db, ids)

    return {"words": words, "total": len(words)}


@router.get("/markdown")
async def export_markdown(
    word_ids: str = Query(None, description="逗号分隔的词语ID"),
    db: AsyncSession = Depends(get_db),
):
    """导出为 Markdown"""
    ids = [int(x) for x in word_ids.split(",")] if word_ids else None
    words = await _get_words(db, ids)

    lines = ["# 考公成语词库导出\n"]
    for w in words:
        lines.append(f"## {w['word']}（{w['pinyin']}）\n")
        lines.append(f"**释义**：{w['meaning']}\n")
        if w["source"]:
            lines.append(f"**出处**：{w['source']}\n")
        if w["example"]:
            lines.append(f"**例句**：{w['example']}\n")
        if w["synonym"]:
            lines.append(f"**近义词**：{w['synonym']}\n")
        if w["antonym"]:
            lines.append(f"**反义词**：{w['antonym']}\n")
        if w["memory_tip"]:
            lines.append(f"**记忆技巧**：{w['memory_tip']}\n")
        lines.append("---\n")

    content = "\n".join(lines)
    return StreamingResponse(
        iter([content]),
        media_type="text/markdown",
        headers={"Content-Disposition": "attachment; filename=idioms.md"},
    )


@router.get("/excel")
async def export_excel(
    word_ids: str = Query(None, description="逗号分隔的词语ID"),
    db: AsyncSession = Depends(get_db),
):
    """导出为 Excel (.xlsx)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ids = [int(x) for x in word_ids.split(",")] if word_ids else None
    words = await _get_words(db, ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成语词库"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    headers = ["词语", "拼音", "释义", "出处", "用法", "例句",
               "近义词", "反义词", "易混词", "记忆技巧", "标签",
               "掌握状态", "复习次数", "错误次数", "收藏"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, w in enumerate(words, 2):
        ws.cell(row=row, column=1, value=w["word"])
        ws.cell(row=row, column=2, value=w["pinyin"])
        ws.cell(row=row, column=3, value=w["meaning"])
        ws.cell(row=row, column=4, value=w["source"])
        ws.cell(row=row, column=5, value=w["usage"])
        ws.cell(row=row, column=6, value=w["example"])
        ws.cell(row=row, column=7, value=w["synonym"])
        ws.cell(row=row, column=8, value=w["antonym"])
        ws.cell(row=row, column=9, value=w["confusable"])
        ws.cell(row=row, column=10, value=w["memory_tip"])
        ws.cell(row=row, column=11, value=",".join(w["tags"]) if isinstance(w["tags"], list) else w["tags"])
        ws.cell(row=row, column=12, value="是" if w["is_mastered"] else "否")
        ws.cell(row=row, column=13, value=w["review_count"])
        ws.cell(row=row, column=14, value=w["error_count"])
        ws.cell(row=row, column=15, value="是" if w["is_favorite"] else "否")

    # 自适应列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=idioms.xlsx"},
    )


@router.get("/pdf")
async def export_pdf(
    word_ids: str = Query(None, description="逗号分隔的词语ID"),
    db: AsyncSession = Depends(get_db),
):
    """导出为 PDF"""
    ids = [int(x) for x in word_ids.split(",")] if word_ids else None
    words = await _get_words(db, ids)

    # 简单的 HTML 转 PDF
    html_parts = ["<!DOCTYPE html><html><head><meta charset='utf-8'>",
                  "<style>body{font-family: sans-serif; margin: 20px;}"
                  "h1{color:#1a73e8;} h2{color:#333; border-bottom:1px solid #ddd; padding-bottom:5px;}"
                  ".meta{color:#666;font-size:14px;}</style></head><body>",
                  "<h1>考公成语词库导出</h1>"]

    for w in words:
        html_parts.append(f"<h2>{w['word']} <span class='meta'>({w['pinyin']})</span></h2>")
        html_parts.append(f"<p><strong>释义：</strong>{w['meaning']}</p>")
        if w["source"]:
            html_parts.append(f"<p><strong>出处：</strong>{w['source']}</p>")
        if w["example"]:
            html_parts.append(f"<p><strong>例句：</strong>{w['example']}</p>")
        if w["synonym"]:
            html_parts.append(f"<p><strong>近义词：</strong>{w['synonym']}</p>")
        html_parts.append("<hr>")

    html_parts.append("</body></html>")
    html_content = "".join(html_parts)

    # 尝试使用 weasyprint 导出 PDF
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=idioms.pdf"},
        )
    except ImportError:
        # weasyprint 可能不易安装，降级返回 HTML
        return StreamingResponse(
            iter([html_content]),
            media_type="text/html",
            headers={"Content-Disposition": "attachment; filename=idioms.html"},
        )
