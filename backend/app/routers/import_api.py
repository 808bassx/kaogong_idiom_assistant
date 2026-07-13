"""
数据导入 API 路由
支持 CSV、JSON、Excel 格式
"""
import csv
import json
import io
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.database import get_db
from app.models import Word
from app.services.review_service import review_service

router = APIRouter(prefix="/api/import", tags=["数据导入"])


@router.post("/csv")
async def import_csv(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """从 CSV 文件导入"""
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="请上传 CSV 文件")

    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    success = 0
    failed = 0
    errors = []

    for row_num, row in enumerate(reader, 2):
        try:
            word_name = row.get("词语", "").strip()
            if not word_name:
                failed += 1
                errors.append(f"第{row_num}行：词语为空")
                continue

            # 检查是否已存在
            result = await db.execute(select(Word).where(Word.word == word_name))
            existing = result.scalar_one_or_none()
            if existing:
                # 更新
                for key, field in [("拼音", "pinyin"), ("释义", "meaning"),
                                    ("出处", "source"), ("用法", "usage"),
                                    ("例句", "example"), ("近义词", "synonym"),
                                    ("反义词", "antonym"), ("易混词", "confusable"),
                                    ("记忆技巧", "memory_tip")]:
                    if key in row and row[key].strip():
                        setattr(existing, field, row[key].strip())
            else:
                word = Word(
                    word=word_name,
                    pinyin=row.get("拼音", "").strip(),
                    meaning=row.get("释义", "").strip(),
                    source=row.get("出处", "").strip(),
                    usage=row.get("用法", "").strip(),
                    example=row.get("例句", "").strip(),
                    synonym=row.get("近义词", "").strip(),
                    antonym=row.get("反义词", "").strip(),
                    confusable=row.get("易混词", "").strip(),
                    memory_tip=row.get("记忆技巧", "").strip(),
                )
                db.add(word)
                await db.flush()
                await review_service.auto_schedule(word.id, word.word, db)

            success += 1
        except Exception as e:
            failed += 1
            errors.append(f"第{row_num}行：{str(e)}")

    await db.flush()
    return {"total": success + failed, "success": success, "failed": failed, "errors": errors[:20]}


@router.post("/json")
async def import_json(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """从 JSON 文件导入"""
    if not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="请上传 JSON 文件")

    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="JSON 格式错误")

    words_data = data if isinstance(data, list) else data.get("words", [])

    success = 0
    failed = 0
    errors = []

    for idx, item in enumerate(words_data):
        try:
            word_name = item.get("word", "").strip()
            if not word_name:
                failed += 1
                errors.append(f"第{idx+1}项：词语为空")
                continue

            result = await db.execute(select(Word).where(Word.word == word_name))
            existing = result.scalar_one_or_none()
            if existing:
                for key in ["pinyin", "meaning", "source", "usage", "example",
                            "synonym", "antonym", "confusable", "memory_tip"]:
                    if key in item and item[key]:
                        setattr(existing, key, item[key])
            else:
                word = Word(
                    word=word_name,
                    pinyin=item.get("pinyin", ""),
                    meaning=item.get("meaning", ""),
                    source=item.get("source", ""),
                    usage=item.get("usage", ""),
                    example=item.get("example", ""),
                    synonym=item.get("synonym", ""),
                    antonym=item.get("antonym", ""),
                    confusable=item.get("confusable", ""),
                    memory_tip=item.get("memory_tip", ""),
                )
                db.add(word)
                await db.flush()
                await review_service.auto_schedule(word.id, word.word, db)

            success += 1
        except Exception as e:
            failed += 1
            errors.append(f"第{idx+1}项：{str(e)}")

    await db.flush()
    return {"total": success + failed, "success": success, "failed": failed, "errors": errors[:20]}


@router.post("/excel")
async def import_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """从 Excel 文件导入 (.xlsx)"""
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # 读取表头
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            headers[col] = str(header_value).strip()

    success = 0
    failed = 0
    errors = []

    for row in range(2, ws.max_row + 1):
        try:
            word_name = ws.cell(row=row, column=1).value
            if not word_name:
                continue
            word_name = str(word_name).strip()

            word_data = {"word": word_name}
            field_map = {"拼音": "pinyin", "释义": "meaning", "出处": "source",
                        "用法": "usage", "例句": "example", "近义词": "synonym",
                        "反义词": "antonym", "易混词": "confusable", "记忆技巧": "memory_tip"}

            for col, header in headers.items():
                if header in field_map:
                    val = ws.cell(row=row, column=col).value
                    if val:
                        word_data[field_map[header]] = str(val).strip()

            result = await db.execute(select(Word).where(Word.word == word_name))
            existing = result.scalar_one_or_none()
            if existing:
                for key, val in word_data.items():
                    if key != "word" and val:
                        setattr(existing, key, val)
            else:
                word = Word(**word_data)
                db.add(word)
                await db.flush()
                await review_service.auto_schedule(word.id, word.word, db)

            success += 1
        except Exception as e:
            failed += 1
            errors.append(f"第{row}行：{str(e)}")

    await db.flush()
    return {"total": success + failed, "success": success, "failed": failed, "errors": errors[:20]}
